# This is a sample Python script.

# Press ⌃R to execute it or replace it with your code.
# Press Double ⇧ to search everywhere for classes, files, tool windows, actions, and settings.


# def print_hi(name):
#     # Use a breakpoint in the code line below to debug your script.
#     print(f'Hi, {name}')  # Press ⌘F8 to toggle the breakpoint.


# Press the green button in the gutter to run the script.
# if __name__ == '__main__':
#     print_hi('PyCharm')

# See PyCharm help at https://www.jetbrains.com/help/pycharm/


# 0616
i = 1
while i <= 5:
    print('*' * i)
    i = i + 1
print("Done")

i = 1
while i <= 5:
    print(i)
    i = i + 1
print("Done")

# 0616  猜谜游戏（while循环）
secret_number = 9
guess_count = 0
guess_limit = 3
while guess_count < guess_limit:
    guess = int(input("Guess: "))
    guess_count += 1
    if guess == secret_number:
        print("You won!")
        break
else:
    print("Sorry, you failed!")

# 0616 哥哥是猪猪

# 0616 汽车游戏（while循环）
command = ""
started = False
while True:
    command = input("> ").lower()
    if command == "start":
        if started:
            print("Car is already started!")
        else:
            started = True
            print("Car started...")
    elif command == "stop":
        if not started:
            print("Car is already stopped!")
        else:
            started = False
            print("Car stopped.")
    elif command == "help":
        print("""
start - to start the car  
stop - to stop the car
quite - to quite    
        """)
    elif command == "quit":
        break
    else:
        print("I don't understand that...")


# 0622 计算商品总价（for循环）
prices = [10,20,30]
total = 0
for price in prices:
    total += price
print(f"Total: {total}")


#0622 嵌套循环（for循环）
numbers = [5, 2, 5, 2, 2]
for i in numbers:
    print("*" * i)

numbers = [5, 2, 5, 2, 2]
for i in numbers:
    output = ""
    for count in range(i):
        output += "X"
    print(output)

#0622 找出最大的数字
number = [1, 2, 8, 3, 4, 5, 6]
print(max(number))

number = [1, 2, 9, 3, 4, 5, 6]
ma = number[0]
for num in number:
    if num > ma:
        ma = num
print(ma)

#0622 矩形数字数组
matrix = [
    [1, 2, 3],
    [4, 5, 6],
    [7, 8, 9]
]
print(matrix[1][2])
for row in matrix:
    for item in row:
        print(item)


#0622 删除列表上的副本
numb = [2, 2, 4, 6, 3, 4, 6, 1]
uni = []
for nu in numb:
    if nu not in uni:
        uni.append(nu)
print(uni)

# 0629 元组
numbers = (1, 2, 3)
print(numbers[0])

# 0629 拆包
coordinates = (1, 2, 3)
x, y, z = coordinates
print(x)

# x, y, z = coordinates相当于下面三行，实现相同结果的简写形式
# x = coordinates[0]
# y = coordinates[1]
# z = coordinates[2]

# 0629 字典
# 使用[]或者调用.get可以防止在用户输入某个字符不是我们字典的一部分时，我们的程序不会报错

customer = {
    "name": "John Smith",
    "age": 30,
    "is_verified": True
}
customer["name"] = "Jack Smith"
print(customer["name"])
print(customer.get("bir", "Octorber 1 1992"))

# 0629 数字转化为英文
phone = input("Phone: ")
digits_mapping = {
    "1": "One",
    "2": "Two",
    "3": "Three",
    "4": "Four"
}
output = ""
for ch in phone:
    output += digits_mapping.get(ch,"!") + " "
print(output)


# 0629 符号转化为emoji表情
message = input(">")
words = message.split(" ")
emojis = {
    ":)": "😊",
    ":(": "😭"
}
output = ""
for word in words:
    output += emojis.get(word,word) + " "
print(output)

# 0629 调用函数
def greet_user():
    print("Hi there!")
    print("Welcome aboard")


print("Start")
greet_user()
print('Finish')

# 0629 参数/使用关键字参数在位置参数之后
def greet_user(fist_name,last_name):
    print(f"Hi {fist_name} {last_name}!")
    print("Welcome aboard")


print("Start")
greet_user("John","Smith")
greet_user("Mary",last_name="Moli")
print('Finish')

# 0629 返回语句
def square(number):
    return number * number


print(square(3))

# 自己整数输入改良版本
def square(number):
    return number * number


number = int(input("Square: "))
print(square(number))


# 0629 创建可重用函数
def emoji_converter(message):
    words = message.split(" ")
    emojis = {
        ":)": "😊",
        ":(": "😭"
    }
    output = ""
    for word in words:
        output += emojis.get(word, word) + " "
    return output


message = input(">")
print(emoji_converter(message))

# 0629 处理系统错误
try:
    age = int(input("Age:"))
    income =20000
    risk = income / age
    print(age)
except ZeroDivisionError:
    print("Age cannot be 0.")
except ValueError:
    print("Invalid value")

# 0629 类
class Point:
    def move(self):
        print("move")

    def draw(self):
        print("draw")


point1 = Point()
point1.x = 10
point1.y = 20
print(point1.x)
point1.draw()


class Point:
    def __init__(self, x, y):
        self.x = x
        self.y = y

    def move(self):
        print("move")

    def draw(self):
        print("draw")


point = Point(11, 20)
print(point.x)


# person新类型
class Person:
    def talk(self):
        print("talk")


john = Person()
john.talk()


class Person:
    def __init__(self, name):
        self.name = name
    def talk(self):
        print("talk")


john = Person("John Smith")
print(john.name)
john.talk()


class Person:
    def __init__(self, name):
        self.name = name

    def talk(self):
        print(f"Hi, I am {self.name}")


john = Person("John Smith")
john.talk()

bob = Person("Bob Smith")
bob.talk()

# 0706 继承
# class Dog:
#    def walk(self):
#       print("walk")


# class Cat:
#     def walk(self):
#         print("walk")


# ————————————————————

class Mammal:
    def walk(self):
        print("walk")


class Cat(Mammal):
    def be_annoying(self):
        print("annoying")


class Dog(Mammal):
    def bark(self):
        print("bark")


cat1 = Cat()
cat1.be_annoying()

dog1 = Dog()
dog1.walk()


# 0706 模块

import converters
from converters import kg_to_lbs

print(converters.kg_to_lbs(70))

print(kg_to_lbs(100))

# 0706 模块
from b_maximum import find_max

numbers = [10, 3, 6, 2]
maximum = find_max(numbers)
print(maximum)

# 0707包，方法一
import ecommerce.shopping
ecommerce.shopping.calc_shopping()

# 0707包，方法二
from ecommerce.shopping import calc_shopping

calc_shopping()

# 0707包，方法三
from ecommerce import shopping

shopping.calc_shopping()


# 0707生成随机数 0～1之间的随机数
import random

for i in range(3):
    print(random.random())

# 10～20以内随机数
import random

for i in range(3):
    print(random.randint(10, 20))

# 随机产生一个领导
import random

members = ["John", "Mary", "Bob", "Mosh"]
leader = random.choice(members)
print(leader)

# Dice = ["1", "2", "3", "4", "5", "6"]
# roll =random.choice(Dice + Dice)
# print(roll)

# 随机骰子
class Dice:
    def roll(self):
        first = random.randint(1, 6)
        second = random.randint(1, 6)
        return first, second


dice = Dice()
print(dice.roll())


# 0707使用目录
# print(path.mkdir())——新建目录
# rmdir()——删除目录
from pathlib import Path

# Absolute path
# c:\Program Files\Microsoft
# /usr/local/bin
# Relative path

from pathlib import Path

path = Path("ecommerce")
print(path.exists())

path = Path()
print(path.glob("*.py"))

path = Path()
for file in path.glob("*"):
    print(file)

# 0709 项目1
import openpyxl as xl
wb = xl.load_workbook("transactions.xlsx")
# print(wb.sheetnames)
# sheet = wb['Sheet1'] 用'Sheet1' 报错, 改成了 '工作表 1'
sheet = wb['工作表 1']
cell = sheet["a1"]
cell = sheet.cell(1, 1)
print(cell.value)

for row in range(1, sheet.max_row + 1):
    print(row)

for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 3)
    print(cell.value)


import openpyxl as xl
from openpyxl.chart import BarChart, Reference

wb = xl.load_workbook("transactions.xlsx")
sheet = wb['工作表 1']
cell = sheet["a1"]
cell = sheet.cell(1, 1)

for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 3)
    corrected_price = cell.value * 2
    corrected_price_cell =sheet.cell(row, 4)
    corrected_price_cell.value = corrected_price

values = Reference(sheet,
          min_row=2,
          max_row=sheet.max_row,
          min_col=4,
          max_col=4)

chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart, "e2")

wb.save("transactions2.xlsx")


# 最终版本
import openpyxl as xl
from openpyxl.chart import BarChart, Reference


def process_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['工作表 1']

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 2
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

    values = Reference(sheet,
                       min_row=2,
                       max_row=sheet.max_row,
                       min_col=4,
                       max_col=4)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, "e2")

    wb.save(filename)
